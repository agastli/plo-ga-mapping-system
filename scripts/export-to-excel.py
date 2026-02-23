#!/usr/bin/env python3
"""
Export PLO-GA mapping data to Excel format
Generates a professional spreadsheet matching PDF format
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# Qatar University Colors
QU_MAROON = "8B1538"
QU_GOLD = "D4AF37"
LIGHT_GOLD = "C8A882"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"
DARK_GRAY = "555555"

def create_mapping_excel(data):
    """Create Excel workbook with professional styling matching PDF"""
    wb = Workbook()
    
    # ===== Sheet 1: Program Information =====
    ws_info = wb.active
    ws_info.title = "Program Info"
    
    current_row = 1
    
    # Add QU logo at the top (centered)
    if data.get('logo_path'):
        try:
            img = XLImage(data['logo_path'])
            # Preserve aspect ratio - QU logo is 4.67:1 (2048x439 pixels)
            img.width = 280  # Increased width for better visibility
            img.height = 60  # 280 / 4.67 = 60 (maintains aspect ratio)
            # Center the logo by placing it in column B
            ws_info.add_image(img, 'B1')
            current_row = 6  # Skip rows for logo
        except Exception as e:
            print(f"Warning: Could not add logo: {e}", file=sys.stderr)
    
    # Add "Academic Planning & Quality Assurance Office" under logo
    ws_info.cell(current_row, 1, 'Academic Planning & Quality Assurance Office')
    ws_info.cell(current_row, 1).font = Font(italic=True, color="808080", size=10)
    ws_info.cell(current_row, 1).alignment = Alignment(horizontal='center')
    ws_info.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 2
    
    # Add decorative title
    ws_info.cell(current_row, 1, data['program_name'])
    ws_info.cell(current_row, 1).font = Font(bold=True, size=18, color=QU_MAROON)
    ws_info.cell(current_row, 1).alignment = Alignment(horizontal='center')
    ws_info.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    ws_info.cell(current_row, 1, 'Program Learning Outcomes to Graduate Attributes Mapping')
    ws_info.cell(current_row, 1).font = Font(size=12, color=DARK_GRAY)
    ws_info.cell(current_row, 1).alignment = Alignment(horizontal='center')
    ws_info.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 2
    
    # Add program information
    info_data = [
        ('College:', data['college_name']),
        ('Department:', data['department_name']),
        ('Language:', data['language']),
        ('Last Updated:', data.get('last_updated', 'N/A'))
    ]
    
    info_start_row = current_row
    for label, value in info_data:
        ws_info.cell(current_row, 1, label)
        ws_info.cell(current_row, 1).font = Font(bold=True, color=QU_MAROON, size=11)
        ws_info.cell(current_row, 1).fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
        
        ws_info.cell(current_row, 2, value)
        ws_info.cell(current_row, 2).font = Font(size=11, color="333333")
        ws_info.cell(current_row, 2).fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin', color=MEDIUM_GRAY),
            right=Side(style='thin', color=MEDIUM_GRAY),
            top=Side(style='thin', color=MEDIUM_GRAY),
            bottom=Side(style='thin', color=MEDIUM_GRAY)
        )
        ws_info.cell(current_row, 1).border = thin_border
        ws_info.cell(current_row, 2).border = thin_border
        
        current_row += 1
    
    # Adjust column widths
    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 75
    
    # Set row heights for logo area
    for row in range(1, 6):
        ws_info.row_dimensions[row].height = 20
    
    # ===== Sheet 2: PLOs =====
    ws_plos = wb.create_sheet("PLOs")
    
    # Header
    ws_plos['A1'] = 'PLO Code'
    ws_plos['B1'] = 'Description'
    
    header_fill = PatternFill(start_color=QU_MAROON, end_color=QU_MAROON, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ['A1', 'B1']:
        ws_plos[cell].fill = header_fill
        ws_plos[cell].font = header_font
        ws_plos[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add PLOs
    for idx, plo in enumerate(data['plos'], start=2):
        ws_plos[f'A{idx}'] = plo['code']
        ws_plos[f'A{idx}'].font = Font(bold=True, color=QU_MAROON)
        ws_plos[f'A{idx}'].alignment = Alignment(vertical='top')
        
        ws_plos[f'B{idx}'] = plo['description']
        ws_plos[f'B{idx}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws_plos.row_dimensions[idx].height = 40  # Adjust row height for wrapped text
    
    # Adjust column widths
    ws_plos.column_dimensions['A'].width = 15
    ws_plos.column_dimensions['B'].width = 100
    
    # ===== Sheet 3: Mapping Matrix (Transposed) =====
    ws_matrix = wb.create_sheet("Mapping Matrix")
    
    # Build transposed matrix: PLOs as columns, competencies as rows with GA headers
    row_idx = 1
    
    # Header row
    ws_matrix.cell(row_idx, 1, 'Graduate Attributes')
    ws_matrix.cell(row_idx, 2, 'Supporting Competencies')
    
    for plo_idx, plo in enumerate(data['plos']):
        ws_matrix.cell(row_idx, 3 + plo_idx, plo['code'])
    
    # Style header row
    for col_idx in range(1, 3 + len(data['plos'])):
        cell = ws_matrix.cell(row_idx, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    row_idx += 1
    
    # Data rows with GA section headers
    for ga in data['gas']:
        # GA section row
        ga_cell = ws_matrix.cell(row_idx, 1)
        ga_cell.value = f"{ga['code']}: {ga['name']}"
        
        # Merge first two columns for GA header
        ws_matrix.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        
        # Style GA row
        ga_cell.fill = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")
        ga_cell.font = Font(bold=True, color=QU_MAROON, size=10)
        ga_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Empty cells for PLO columns in GA row
        for plo_idx in range(len(data['plos'])):
            plo_cell = ws_matrix.cell(row_idx, 3 + plo_idx)
            plo_cell.fill = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")
        
        row_idx += 1
        
        # Competency rows under this GA
        for comp in ga['competencies']:
            # Empty GA column
            ws_matrix.cell(row_idx, 1).value = ''
            ws_matrix.cell(row_idx, 1).fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
            
            # Competency name
            comp_cell = ws_matrix.cell(row_idx, 2)
            comp_cell.value = f"{comp['code']} – {comp['name']}"
            comp_cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
            comp_cell.font = Font(size=9)
            comp_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Weights for each PLO
            for plo_idx, plo in enumerate(data['plos']):
                weight = plo['mappings'].get(comp['code'], '0.00')
                weight_cell = ws_matrix.cell(row_idx, 3 + plo_idx)
                weight_cell.value = weight
                weight_cell.alignment = Alignment(horizontal='center', vertical='center')
                weight_cell.font = Font(size=9)
            
            row_idx += 1
    
    # Add borders to all cells in matrix
    thin_border = Border(
        left=Side(style='thin', color='808080'),
        right=Side(style='thin', color='808080'),
        top=Side(style='thin', color='808080'),
        bottom=Side(style='thin', color='808080')
    )
    
    for row in ws_matrix.iter_rows(min_row=1, max_row=row_idx-1, min_col=1, max_col=2 + len(data['plos'])):
        for cell in row:
            cell.border = thin_border
    
    # Adjust column widths
    ws_matrix.column_dimensions['A'].width = 20
    ws_matrix.column_dimensions['B'].width = 45
    for plo_idx in range(len(data['plos'])):
        ws_matrix.column_dimensions[get_column_letter(3 + plo_idx)].width = 10
    
    # Add summary below matrix
    row_idx += 2
    total_comps = sum(len(ga['competencies']) for ga in data['gas'])
    summary_text = f"Total Mappings: {data['total_mappings']} | Total PLOs: {len(data['plos'])} | Total Competencies: {total_comps}"
    ws_matrix.cell(row_idx, 1, summary_text)
    ws_matrix.cell(row_idx, 1).font = Font(bold=True, size=10)
    ws_matrix.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2 + len(data['plos']))
    
    # ===== Sheet 4: Justifications =====
    ws_just = wb.create_sheet("Justifications")
    
    # Header
    ws_just['A1'] = 'Code'
    ws_just['B1'] = 'Competency Name'
    ws_just['C1'] = 'Justification'
    
    for cell in ['A1', 'B1', 'C1']:
        ws_just[cell].fill = header_fill
        ws_just[cell].font = header_font
        ws_just[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add justifications
    for idx, just in enumerate(data['justifications'], start=2):
        # Code
        ws_just[f'A{idx}'] = just['competency_code']
        ws_just[f'A{idx}'].fill = PatternFill(start_color=QU_MAROON, end_color=QU_MAROON, fill_type="solid")
        ws_just[f'A{idx}'].font = Font(bold=True, color="FFFFFF", size=9)
        ws_just[f'A{idx}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Name
        ws_just[f'B{idx}'] = just['competency_name']
        ws_just[f'B{idx}'].fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
        ws_just[f'B{idx}'].font = Font(bold=True, color=QU_MAROON, size=9)
        ws_just[f'B{idx}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Justification text
        ws_just[f'C{idx}'] = just['text']
        ws_just[f'C{idx}'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='justify')
        ws_just[f'C{idx}'].font = Font(size=9, color="333333")
        
        # Add borders
        for col in ['A', 'B', 'C']:
            ws_just[f'{col}{idx}'].border = thin_border
        
        # Adjust row height for wrapped text
        ws_just.row_dimensions[idx].height = 60
    
    # Adjust column widths
    ws_just.column_dimensions['A'].width = 10
    ws_just.column_dimensions['B'].width = 35
    ws_just.column_dimensions['C'].width = 90
    
    return wb

def main():
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'No input data file provided'}))
        sys.exit(1)
    
    try:
        # Read JSON data from file (same as PDF export approach)
        data_file_path = sys.argv[1]
        with open(data_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Create workbook
        wb = create_mapping_excel(data)
        
        # Save to output path
        import os
        output_path = data.get('output_path', os.path.join(os.path.dirname(os.path.dirname(__file__)), 'temp', 'mapping_output.xlsx'))
        wb.save(output_path)
        
        print(json.dumps({'success': True, 'output_path': output_path}))
    
    except Exception as e:
        import traceback
        print(json.dumps({'error': str(e), 'traceback': traceback.format_exc()}))
        sys.exit(1)

if __name__ == '__main__':
    main()
