#!/usr/bin/env python3
"""
Export analytics data to Excel format
"""

import sys
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

def create_analytics_excel(data, output_path, logo_path):
    """Generate Excel analytics report"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics Report"
    
    # QU Colors
    maroon_fill = PatternFill(start_color="8B1538", end_color="8B1538", fill_type="solid")
    gold_fill = PatternFill(start_color="CFB87C", end_color="CFB87C", fill_type="solid")
    light_gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    white_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(color="8B1538", bold=True, size=18)
    heading_font = Font(color="8B1538", bold=True, size=14)
    
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    current_row = 1
    
    # Add logo if available
    if logo_path and os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            # Preserve aspect ratio
            img.width = 200
            img.height = int(200 / 4.67)  # 4.67:1 aspect ratio
            ws.add_image(img, f'A{current_row}')
            current_row += 3
        except Exception as e:
            print(f"Warning: Could not add logo: {e}", file=sys.stderr)
    
    # Office name
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "Academic Planning & Quality Assurance Office"
    cell.font = Font(size=11, color="666666")
    cell.alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Title
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = data.get('title', 'Analytics Report')
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Metrics section
    if 'metrics' in data and len(data['metrics']) > 0:
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "Key Metrics"
        cell.font = heading_font
        current_row += 1
        
        # Metrics table header
        ws[f'A{current_row}'] = "Metric"
        ws[f'B{current_row}'] = "Value"
        for col in ['A', 'B']:
            cell = ws[f'{col}{current_row}']
            cell.fill = maroon_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        current_row += 1
        
        # Metrics data
        for metric in data['metrics']:
            ws[f'A{current_row}'] = metric['label']
            ws[f'B{current_row}'] = metric['value']
            for col in ['A', 'B']:
                cell = ws[f'{col}{current_row}']
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            current_row += 1
        
        current_row += 2
    
    # Data table
    if 'table_data' in data and len(data['table_data']) > 0:
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "Detailed Data"
        cell.font = heading_font
        current_row += 1
        
        table_data = data['table_data']
        
        # Header row
        for col_idx, header in enumerate(table_data[0], start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = maroon_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        current_row += 1
        
        # Data rows
        for row_data in table_data[1:]:
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # Alternate row colors
                if current_row % 2 == 0:
                    cell.fill = light_gray_fill
            current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_path)
    
    return {"success": True, "path": output_path}

if __name__ == "__main__":
    try:
        if len(sys.argv) < 2:
            print(json.dumps({"success": False, "error": "No input file provided"}))
            sys.exit(1)
        
        temp_file = sys.argv[1]
        with open(temp_file, 'r', encoding='utf-8') as f:
            input_data = json.load(f)
        
        data = input_data['data']
        output_path = input_data['output_path']
        logo_path = input_data.get('logo_path', '')
        
        result = create_analytics_excel(data, output_path, logo_path)
        print(json.dumps(result))
        
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}))
        sys.exit(1)
