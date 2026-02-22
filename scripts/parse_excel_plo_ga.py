#!/usr/bin/env python3
"""
Excel PLO-GA Mapping Parser
Parses Excel files containing PLO to Graduate Attribute mappings
Supports both English and Arabic content
"""

import openpyxl
import sys
import json
import re

def parse_plo_mappings(mapping_text):
    """
    Parse PLO mappings from text like:
    - "PLO6 (1.00)"
    - "PLO4 (0.51) & PLO2 (0.49)"
    - "-"
    
    Returns: List of {plo_number, weight} dicts
    """
    if not mapping_text or mapping_text.strip() == "-":
        return []
    
    mappings = []
    # Pattern: PLO followed by number, then weight in parentheses
    pattern = r'PLO(\d+)\s*\(([0-9.]+)\)'
    matches = re.findall(pattern, str(mapping_text))
    
    for plo_num, weight in matches:
        mappings.append({
            "plo_number": int(plo_num),
            "weight": float(weight)
        })
    
    return mappings

def detect_language(text):
    """
    Detect if text is Arabic or English based on character ranges
    """
    if not text:
        return "en"
    
    arabic_chars = sum(1 for c in text if '\u0600' <= c <= '\u06FF' or '\u0750' <= c <= '\u077F')
    return "ar" if arabic_chars > len(text) * 0.3 else "en"

def parse_excel_plo_ga(filepath):
    """
    Parse Excel file containing PLO-GA mappings
    
    Expected structure:
    - Sheet "Program Information": Contains program name, college, department
    - Sheet "Justifications": Contains competency mappings and justifications
    - Column A: Competency Code (e.g., C1-1, GA1)
    - Column B: Competency Name
    - Column C: Mapped PLOs (e.g., "PLO6 (1.00)" or "PLO4 (0.51) & PLO2 (0.49)")
    - Column D: Justification text
    
    Returns: Dict with plos, mappings, and justifications for database import
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Extract program information from "Program Information" sheet
    program_name = None
    college_name = None
    department_name = None
    language = "en"
    
    if "Program Information" in wb.sheetnames:
        info_ws = wb["Program Information"]
        for row in info_ws.iter_rows(min_row=2, max_row=10, values_only=True):
            if not row[0]:
                continue
            field = str(row[0]).strip()
            value = str(row[1]).strip() if row[1] else None
            
            if field == "Program Name" and value:
                program_name = value
            elif field == "College" and value:
                college_name = value
            elif field == "Department" and value:
                department_name = value
            elif field == "Language" and value:
                language = "ar" if "arabic" in value.lower() else "en"
    
    # Find "Justifications" sheet for mappings
    if "Justifications" in wb.sheetnames:
        ws = wb["Justifications"]
    elif "Mapping" in wb.sheetnames:
        ws = wb["Mapping"]
    else:
        # Skip "Program Information" sheet if it exists
        for sheet_name in wb.sheetnames:
            if sheet_name != "Program Information":
                ws = wb[sheet_name]
                break
    
    # Collect all unique PLOs mentioned in the document
    plo_set = set()
    mappings_list = []
    justifications_list = []
    current_ga = None
    
    # Start from row 2 (skip header)
    for row in ws.iter_rows(min_row=2, values_only=True):
        comp_code = row[0]  # Column A
        comp_name = row[1]  # Column B
        plo_mapping = row[2] if len(row) > 2 else None  # Column C
        justification = row[3] if len(row) > 3 else None  # Column D
        
        # Skip empty rows
        if not comp_code:
            continue
        
        comp_code = str(comp_code).strip()
        
        # Check if this is a GA header row (e.g., "GA1", "GA2")
        if comp_code.startswith("GA") and not "-" in comp_code:
            current_ga = comp_code
            continue
        
        # Check if this is a competency row (e.g., "C1-1", "C2-3")
        if comp_code.startswith("C") and "-" in comp_code:
            # Parse PLO mappings
            plo_mappings = parse_plo_mappings(plo_mapping) if plo_mapping else []
            
            # Add PLOs to set and create mapping entries
            for plo_map in plo_mappings:
                plo_code = f"PLO{plo_map['plo_number']}"
                plo_set.add(plo_map['plo_number'])
                
                mappings_list.append({
                    "ploCode": plo_code,
                    "competencyCode": comp_code,
                    "weight": plo_map['weight']
                })
            
            # Add justification
            if justification:
                just_text = str(justification).strip()
                # Clean up Excel line breaks
                just_text = just_text.replace('_x000D_', '').replace('\r', '').replace('\n\n\n', '\n\n').strip()
                lang = detect_language(just_text)
                
                justifications_list.append({
                    "gaCode": current_ga,
                    "competencyCode": comp_code,
                    "textEn": just_text if lang == "en" else None,
                    "textAr": just_text if lang == "ar" else None
                })
    
    # Create PLO list
    plos_list = []
    for plo_num in sorted(plo_set):
        plos_list.append({
            "code": f"PLO{plo_num}",
            "descriptionEn": None,  # Not available in Excel
            "descriptionAr": None,  # Not available in Excel
            "sortOrder": plo_num
        })
    
    # Detect document language from justifications
    doc_language = "en"
    if justifications_list:
        # Check first justification
        first_just = justifications_list[0]
        if first_just.get("textAr"):
            doc_language = "ar"
    
    return {
        "success": True,
        "data": {
            "programInfo": {
                "programNameEn": program_name or "Extracted from Excel",
                "departmentEn": department_name or "See selection",
                "collegeEn": college_name or "See selection",
                "language": language if language != "en" else doc_language
            },
            "plos": plos_list,
            "mappings": mappings_list,
            "justifications": justifications_list
        }
    }

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 parse_excel_plo_ga.py <excel_file>", file=sys.stderr)
        sys.exit(1)
    
    filepath = sys.argv[1]
    
    try:
        result = parse_excel_plo_ga(filepath)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}), file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
